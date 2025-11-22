# export_utils.py
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from io import BytesIO

COLUMN_LABELS = {
    "nom": {"fr": "Nom", "ar": "الاسم"},
    "prenoms": {"fr": "Prénoms", "ar": "الأسماء"},
    "grade": {"fr": "Grade", "ar": "الرتبة"},
    "specialite": {"fr": "Spécialité", "ar": "التخصص"},
    "ecole_origine": {"fr": "École d'origine", "ar": "المدرسة الأصلية"},
    "cin": {"fr": "CIN", "ar": "بطاقة الهوية"},
    "matricule": {"fr": "Matricule", "ar": "الرقم"},
    "telephone": {"fr": "N° téléphone", "ar": "الهاتف"},
    "email": {"fr": "Adresse e-mail", "ar": "البريد الإلكتروني"},
    "date_affectation": {"fr": "Date d’affectation", "ar": "تاريخ التعيين"},
    "date_passage_grade": {"fr": "Date passage grade", "ar": "تاريخ الترقية"},
    "cv": {"fr": "CV", "ar": "السيرة الذاتية"},
    "pv_affectation": {"fr": "PV Affectation", "ar": "محضر التعيين"},
    "decret_officiel": {"fr": "Décret Officiel", "ar": "المرسوم الرسمي"},
    "fiche_fonction": {"fr": "Fiche Fonction", "ar": "بطاقة الوظيفة"},
    "fiche_module_fr": {"fr": "Fiche Module (FR)", "ar": "بطاقة الوحدة (فر)"},
    "fiche_module_en": {"fr": "Fiche Module (EN)", "ar": "بطاقة الوحدة (إن)"},
}

def generate_excel(personnels, columns, lang='fr', base_url='http://127.0.0.1:8000'):
    """
    Génère un Excel en mémoire contenant les colonnes demandées.
    Pour les FileField (cv, pv_affectation, ...), crée un hyperlien
    si le fichier est présent. Retourne un BytesIO prêt à être renvoyé.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Personnels"

    # En-têtes traductions
    headers = [COLUMN_LABELS.get(col, {}).get(lang, col) for col in columns]
    ws.append(headers)

    # Parcours des personnels et remplissage
    for p in personnels:
        row_values = []
        for col in columns:
            # Récupère l'attribut, sans appeler .url directement
            file_field_candidate = getattr(p, col, None)

            # Si c'est potentiellement un FileField (django.db.models.fields.files.FieldFile)
            is_filefield = hasattr(file_field_candidate, "name") and bool(file_field_candidate.name)
            if is_filefield:
                # construire url complète de façon sûre
                try:
                    relative_url = file_field_candidate.url  # peut lever ValueError si pas de fichier
                    full_url = f"{base_url}{relative_url}"
                except Exception:
                    # si .url échoue, tombe en back-up (écriture du nom ou vide)
                    full_url = None

                if full_url:
                    # on stocke une valeur spéciale pour traitement après append
                    row_values.append({"type": "file", "url": full_url, "label": "Ouvrir le fichier"})
                else:
                    row_values.append({"type": "text", "value": ""})
            else:
                # colonne ordinaire ou filefield vide
                val = getattr(p, col, "")
                # formater dates si besoin
                try:
                    # si val est date/datetime -> formattage propre
                    if hasattr(val, "strftime"):
                        row_values.append({"type": "text", "value": val.strftime("%d/%m/%Y")})
                    else:
                        row_values.append({"type": "text", "value": str(val) if val is not None else ""})
                except Exception:
                    row_values.append({"type": "text", "value": str(val or "")})

        # Append row with placeholders (strings or dicts)
        ws.append([ (cell["value"] if cell["type"] == "text" else cell["label"]) for cell in row_values ])

        # After append, set hyperlinks for file cells on the last row
        last_row_index = ws.max_row
        for col_idx, cell_data in enumerate(row_values, start=1):
            if cell_data["type"] == "file":
                cell = ws.cell(row=last_row_index, column=col_idx)
                # set hyperlink + style
                cell.hyperlink = cell_data["url"]
                cell.value = cell_data.get("label", "Ouvrir le fichier")
                cell.font = Font(color="0000FF", underline="single")

    # Ajuster largeur colonnes
    for i, col in enumerate(columns, 1):
        column_letter = get_column_letter(i)
        max_length = len(headers[i-1]) if headers[i-1] else 10
        for row in ws.iter_rows(min_row=2, min_col=i, max_col=i):
            v = row[0].value
            if v is not None:
                max_length = max(max_length, len(str(v)))
        ws.column_dimensions[column_letter].width = max_length + 4

    # Retour en mémoire
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer



""""
COLUMN_LABELS = {
    "nom": {"fr": "Nom", "ar": "الاسم"},
    "prenoms": {"fr": "Prénoms", "ar": "الأسماء"},
    "grade": {"fr": "Grade", "ar": "الرتبة"},
    "specialite": {"fr": "Spécialité", "ar": "التخصص"},
    "ecole_origine": {"fr": "École d'origine", "ar": "المدرسة الأصلية"},
    "cin": {"fr": "CIN", "ar": "بطاقة الهوية"},
    "matricule": {"fr": "Matricule", "ar": "الرقم"},
    "telephone": {"fr": "N° téléphone", "ar": "الهاتف"},
    "email": {"fr": "Adresse e-mail", "ar": "البريد الإلكتروني"},
    "date_affectation": {"fr": "Date d’affectation", "ar": "تاريخ التعيين"},
    "date_passage_grade": {"fr": "Date passage grade", "ar": "تاريخ الترقية"},
}

# les imports 
import openpyxl
from openpyxl.utils import get_column_letter
from io import BytesIO

# from reportlab.platypus import Table, TableStyle
# from reportlab.lib import colors

def generate_excel(personnels, columns, lang='fr'):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Personnels"

    # En-têtes
    headers = [COLUMN_LABELS.get(col, {}).get(lang, col) for col in columns]
    ws.append(headers)

    # Lignes
    for p in personnels:
        row = [str(getattr(p, col, "")) for col in columns]
        ws.append(row)


    # Ajuster automatiquement la largeur des colonnes
    for i, col in enumerate(columns, 1):
        column_letter = get_column_letter(i)
        max_length = len(col)

        for row in ws.iter_rows(min_row=2, min_col=i, max_col=i):
            cell_value = row[0].value
            if cell_value:
                max_length = max(max_length, len((str(cell_value))))
        
        ws.column_dimensions[column_letter].width = max_length + 2

    # sauvegarder en mémoire
    buffer = BytesIO()
    
    wb.save(buffer)
    buffer.seek(0)

    return buffer
"""

